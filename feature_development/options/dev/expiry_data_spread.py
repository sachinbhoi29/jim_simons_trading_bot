# fetch_options_and_clean.py
import yfinance as yf
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

def fetch_nifty_options(csv_file_path, save_dir="feature_development/options/dev"):
    symbol = "NIFTY"
    
    # ---------------- Step 1: Get Spot Price ----------------
    print("Fetching current spot price from Yahoo Finance...")
    ticker = yf.Ticker("^NSEI")   # NSE Nifty 50 Index
    spot_price = ticker.history(period="1d")["Close"].iloc[-1]
    spot_price = round(float(spot_price), 2)
    atm_strike = round(spot_price / 50) * 50  # NSE has 50pt strikes
    print(f"Spot Price: {spot_price}, ATM Strike: {atm_strike}")

    # ---------------- Step 2: Read Option Chain from CSV ----------------
    print(f"Reading option chain from {csv_file_path}...")
    df_raw = pd.read_csv(csv_file_path)

    # Get nearest expiry from CSV
    expiry_dates = sorted(df_raw["expiryDate"].unique())
    nearest_expiry = expiry_dates[0]   # first expiry = nearest
    print(f"Using nearest expiry: {nearest_expiry}")

    # ---------------- Step 3: Build Clean Data ----------------
    df_expiry = df_raw[df_raw["expiryDate"] == nearest_expiry].copy()
    df_expiry["spotPrice"] = spot_price  # add spot column

    # Calls
    df_expiry["Call_LTP"] = df_expiry["CE_lastPrice"]
    df_expiry["Call_OI"] = df_expiry["CE_openInterest"]
    df_expiry["Call_IV"] = df_expiry["CE_impliedVolatility"]
    df_expiry["Call_Type"] = df_expiry["strikePrice"].apply(
        lambda x: "ITM" if x < spot_price else ("ATM" if x == atm_strike else "OTM")
    )

    # Puts
    df_expiry["Put_LTP"] = df_expiry["PE_lastPrice"]
    df_expiry["Put_OI"] = df_expiry["PE_openInterest"]
    df_expiry["Put_IV"] = df_expiry["PE_impliedVolatility"]
    df_expiry["Put_Type"] = df_expiry["strikePrice"].apply(
        lambda x: "ITM" if x > spot_price else ("ATM" if x == atm_strike else "OTM")
    )

    # ---------------- Step 4: Keep All Required Columns ----------------
    df = df_expiry[[
        "expiryDate", "spotPrice", "strikePrice",
        "CE_lastPrice", "CE_openInterest", "CE_impliedVolatility",
        "PE_lastPrice", "PE_openInterest", "PE_impliedVolatility",
        "Call_LTP", "Call_OI", "Call_IV", "Call_Type",
        "Put_LTP", "Put_OI", "Put_IV", "Put_Type"
    ]]

    # ---------------- Step 5: Save ----------------
    os.makedirs(save_dir, exist_ok=True)
    base_filename = f"{symbol}_options_{nearest_expiry.replace('-', '')}"
    csv_file = os.path.join(save_dir, f"{base_filename}.csv")
    excel_file = os.path.join(save_dir, f"{base_filename}.xlsx")

    df.to_csv(csv_file, index=False)
    df.to_excel(excel_file, index=False)

    # Highlight ITM in Excel
    wb = load_workbook(excel_file)
    ws = wb.active
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for row in range(2, ws.max_row + 1):
        if ws[f"M{row}"].value == "ITM":  # Call_Type col
            for col in ["J", "K", "L", "M"]:
                ws[f"{col}{row}"].fill = green_fill
        if ws[f"Q{row}"].value == "ITM":  # Put_Type col
            for col in ["N", "O", "P", "Q"]:
                ws[f"{col}{row}"].fill = green_fill

    wb.save(excel_file)

    print(f"Options data saved as CSV: {csv_file}")
    print(f"Options data saved with ITM highlighted in Excel: {excel_file}")

    return csv_file, excel_file


# ---------------- Run ----------------
csv_file = "feature_development/options/dev/NIFTY_optionchain_raw.csv"  # your raw CSV file
csv_path, excel_path = fetch_nifty_options(csv_file_path=csv_file)
print("Cleaned files ready:", csv_path, excel_path)
