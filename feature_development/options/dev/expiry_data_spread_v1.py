# fetch_options_and_clean.py
# Validated values from https://www.nseindia.com/option-chain?symbolCode=-10006&symbol=NIFTY&symbol=NIFTY&instrument=-&date=-&segmentLink=17&symbolCount=2&segmentLink=17
import yfinance as yf
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import os
import datetime

def fetch_nifty_options(csv_file_path, save_dir="feature_development/options/dev"):
    symbol = "NIFTY"
    
    # ---------------- Step 1: Get Spot Price ----------------
    print("Fetching current spot price from Yahoo Finance...")
    ticker = yf.Ticker("^NSEI")   # NSE Nifty 50 Index
    spot_price = ticker.history(period="1d")["Close"].iloc[-1]
    spot_price = round(float(spot_price), 2)
    atm_strike = round(spot_price / 50) * 50  # NSE has 50pt strikes
    print(f"Spot Price: {spot_price}, ATM Strike: {atm_strike}")

    # ---------------- Step 2: Read Option Chain from CSV ----------------
    print(f"Reading option chain from {csv_file_path}...")
    df_raw = pd.read_csv(csv_file_path)

    # Convert string to datetime just for sorting/comparison
    expiry_dates_dt = [datetime.datetime.strptime(d, "%d-%b-%Y") for d in df_raw["expiryDate"].unique()]
    today = datetime.datetime.today()

    # Keep only future dates
    future_expiries = [d for d in expiry_dates_dt if d >= today]

    # Sort ascending and take next 7
    next_7_expiries = sorted(future_expiries)[:7]

    # Show options to user
    print("\nNext 7 expiry dates:")
    for i, d in enumerate(next_7_expiries, 1):
        print(f"{i}. {d.strftime('%d-%b-%Y')}")

    # Let user pick one
    choice = int(input("\nEnter the number of expiry you want to process (1-7): "))
    nearest_expiry = next_7_expiries[choice-1].strftime('%d-%b-%Y')
    print(f"\nUsing expiry: {nearest_expiry}")

    # ---------------- Step 3: Filter expiry ----------------
    df_expiry = df_raw[df_raw["expiryDate"] == nearest_expiry].copy()

    df_expiry["spotPrice"] = spot_price  # add spot column

    # ---------------- Step 4: NSE-style Option Chain ----------------
    df = pd.DataFrame({
        # Calls (left side)
        "Call_OI": df_expiry.get("CE_openInterest"),
        "Call_Chng_in_OI": df_expiry.get("CE_changeinOpenInterest"),
        "Call_Volume": df_expiry.get("CE_totalTradedVolume"),
        "Call_IV": df_expiry.get("CE_impliedVolatility"),
        "Call_LTP": df_expiry.get("CE_lastPrice"),
        "Call_Chng": df_expiry.get("CE_change"),
        "Call_BidQty": df_expiry.get("CE_bidQty"),
        "Call_Bid": df_expiry.get("CE_bidprice"),
        "Call_Ask": df_expiry.get("CE_askPrice"),
        "Call_AskQty": df_expiry.get("CE_askQty"),

        # Strike (middle)
        "StrikePrice": df_expiry.get("strikePrice"),

        # Puts (right side)
        "Put_BidQty": df_expiry.get("PE_bidQty"),
        "Put_Bid": df_expiry.get("PE_bidprice"),
        "Put_Ask": df_expiry.get("PE_askPrice"),
        "Put_AskQty": df_expiry.get("PE_askQty"),
        "Put_Chng": df_expiry.get("PE_change"),
        "Put_LTP": df_expiry.get("PE_lastPrice"),
        "Put_IV": df_expiry.get("PE_impliedVolatility"),
        "Put_Volume": df_expiry.get("PE_totalTradedVolume"),
        "Put_Chng_in_OI": df_expiry.get("PE_changeinOpenInterest"),
        "Put_OI": df_expiry.get("PE_openInterest"),
    })

    # ---------------- Step 5: Mark ITM/ATM/OTM ----------------
    # Correct ITM/ATM/OTM classification
    df["Call_Type"] = df["StrikePrice"].apply(
        lambda x: "ITM" if x <= spot_price else "OTM"
    )
    df["Put_Type"] = df["StrikePrice"].apply(
        lambda x: "ITM" if x >= spot_price else "OTM"
    )
    df["ATM"] = df["StrikePrice"].apply(
        lambda x: "ATM" if x == atm_strike else ""
    )

    # ---------------- Step 5b: Append remaining columns after Puts ----------------
    # Keep all columns from raw CSV that are not already in df
    existing_cols = set(df.columns)
    extra_cols = [c for c in df_expiry.columns if c not in existing_cols]

    # Append them **after the Put columns**, with their data
    for col in extra_cols:
        df[col] = df_expiry[col]

    # ---------------- Step 6: Save ----------------
    os.makedirs(save_dir, exist_ok=True)
    base_filename = f"{symbol}_options_{nearest_expiry.replace('-', '')}"
    csv_file = os.path.join(save_dir, f"{base_filename}.csv")
    excel_file = os.path.join(save_dir, f"{base_filename}.xlsx")

    df.to_csv(csv_file, index=False)
    df.to_excel(excel_file, index=False)

    # ---------------- Step 7: Robust Excel Formatting ----------------
    wb = load_workbook(excel_file)
    ws = wb.active

    # Build header -> column index mapping from the first row
    headers = [cell.value for cell in ws[1]]
    header_index = {h: i + 1 for i, h in enumerate(headers)}

    # Required columns for formatting
    required = ["StrikePrice", "Call_Type", "Put_Type", "Call_LTP", "Put_LTP", "Call_Chng", "Put_Chng"]
    missing = [c for c in required if c not in header_index]
    if missing:
        print("Warning: Missing expected columns for Excel formatting:", missing)
        print("Saved files, but skipped Excel formatting. If you expect these columns, check the CSV input.")
        wb.save(excel_file)
        return csv_file, excel_file

    # Get numeric column indices
    strike_col = header_index["StrikePrice"]
    call_type_col = header_index["Call_Type"]
    put_type_col = header_index["Put_Type"]
    call_ltp_col = header_index["Call_LTP"]
    put_ltp_col = header_index["Put_LTP"]
    call_chng_col = header_index["Call_Chng"]
    put_chng_col = header_index["Put_Chng"]

    # Define call/put block ranges relative to strike column
    call_block_start = 1
    call_block_end = strike_col - 1
    put_block_start = strike_col + 1
    first_type_col = min(call_type_col, put_type_col)
    put_block_end = first_type_col - 1

    # Styles
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    bold_font = Font(bold=True)
    green_font = Font(color="008000")   # dark green
    red_font   = Font(color="FF0000")   # red

    # Apply formatting row by row
    for row in range(2, ws.max_row + 1):
        # Highlight ITM calls (left block) only
        if ws.cell(row=row, column=call_type_col).value == "ITM":
            for col in range(call_block_start, call_block_end + 1):
                ws.cell(row=row, column=col).fill = green_fill

        # Highlight ITM puts (right block) only
        if ws.cell(row=row, column=put_type_col).value == "ITM":
            for col in range(put_block_start, put_block_end + 1):
                ws.cell(row=row, column=col).fill = green_fill

        # Bold Strike
        ws.cell(row=row, column=strike_col).font = bold_font

        # Bold LTPs
        ws.cell(row=row, column=call_ltp_col).font = bold_font
        ws.cell(row=row, column=put_ltp_col).font = bold_font

        # Dynamic red/green font for Change columns
        call_chng_val = ws.cell(row=row, column=call_chng_col).value
        put_chng_val = ws.cell(row=row, column=put_chng_col).value

        if isinstance(call_chng_val, (int, float)):
            if call_chng_val > 0:
                ws.cell(row=row, column=call_chng_col).font = green_font
            elif call_chng_val < 0:
                ws.cell(row=row, column=call_chng_col).font = red_font

        if isinstance(put_chng_val, (int, float)):
            if put_chng_val > 0:
                ws.cell(row=row, column=put_chng_col).font = green_font
            elif put_chng_val < 0:
                ws.cell(row=row, column=put_chng_col).font = red_font

    wb.save(excel_file)

    print(f"Options data saved as CSV: {csv_file}")
    print(f"Options data saved with ITM highlighted + formatting in Excel: {excel_file}")

    return csv_file, excel_file


# ---------------- Run ---------------
if __name__ == "__main__":
    csv_file = "feature_development/options/dev/NIFTY_optionchain_raw.csv"  # your raw CSV file
    csv_path, excel_path = fetch_nifty_options(csv_file_path=csv_file)
    print("Cleaned files ready:", csv_path, excel_path)
