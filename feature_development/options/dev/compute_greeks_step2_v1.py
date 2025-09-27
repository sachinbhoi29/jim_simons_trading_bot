# compute_greeks.py
# greeks calculation, they are not exact but are very close, check for direction as well
import os
from datetime import datetime
import numpy as np
import pandas as pd
from scipy.stats import norm
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import PatternFill, Font


def _parse_date_try_formats(s):
    """Try common date formats used in NSE data like '30-Dec-25', '30-Dec-2025', '2025-12-30'."""
    if pd.isna(s):
        return None
    for fmt in ("%d-%b-%y", "%d-%b-%Y", "%d-%m-%y", "%d-%m-%Y", "%Y-%m-%d", "%d-%b-%y"):
        try:
            return datetime.strptime(str(s), fmt).date()
        except Exception:
            continue
    # last resort: pandas
    try:
        return pd.to_datetime(s, dayfirst=True).date()
    except Exception:
        return None

# def black_scholes_greeks(S, K, r, q, sigma, T, option_type):
#     """
#     Returns Delta, Gamma, Vega, Theta (per day), Rho.
#     S: spot
#     K: strike
#     r: risk-free rate (annual, decimal)
#     q: dividend yield (annual, decimal) - for index usually 0
#     sigma: implied volatility (annual, decimal)
#     T: time to expiry in years (float)
#     option_type: 'call' or 'put'
#     """
#     # Handle degenerate cases
#     if sigma is None or sigma <= 0 or T <= 0 or S <= 0 or K <= 0:
#         return (np.nan, np.nan, np.nan, np.nan, np.nan)

#     sqrtT = np.sqrt(T)
#     d1 = (np.log(S / K) + (r - q + 0.5 * sigma ** 2) * T) / (sigma * sqrtT)
#     d2 = d1 - sigma * sqrtT

#     # pdf and cdf
#     pdf_d1 = norm.pdf(d1)
#     cdf_d1 = norm.cdf(d1)
#     cdf_d2 = norm.cdf(d2)

#     # Greeks
#     if option_type.lower().startswith("c"):
#         delta = np.exp(-q * T) * cdf_d1
#         rho = K * T * np.exp(-r * T) * cdf_d2
#         # Theta (annual) for call:
#         theta_annual = (- (S * sigma * np.exp(-q * T) * pdf_d1) / (2 * sqrtT)
#                         - r * K * np.exp(-r * T) * cdf_d2
#                         + q * S * np.exp(-q * T) * cdf_d1)
#     else:
#         # put
#         delta = np.exp(-q * T) * (cdf_d1 - 1)
#         rho = -K * T * np.exp(-r * T) * norm.cdf(-d2)
#         theta_annual = (- (S * sigma * np.exp(-q * T) * pdf_d1) / (2 * sqrtT)
#                         + r * K * np.exp(-r * T) * norm.cdf(-d2)
#                         - q * S * np.exp(-q * T) * norm.cdf(-d1))

#     gamma = (np.exp(-q * T) * pdf_d1) / (S * sigma * sqrtT)
#     vega = (S * np.exp(-q * T) * pdf_d1 * sqrtT) / 100.0  # per 1% vol
#     theta_per_day = theta_annual / 365.0
#     # convert vega to per 1% (optional) — we'll keep as absolute (per 1.0 vol)
#     return (delta, gamma, vega, theta_per_day, rho)
def black_scholes_greeks(S, K, r, q, sigma, T, option_type, zerodha_mode=True):
    """
    Returns Delta, Gamma, Vega, Theta (per day), Rho.
    - If zerodha_mode=True, tweaks formulas to align with Zerodha's convention
      (mainly in Theta, assuming r ≈ 0 so theta is symmetric).
    """
    if sigma is None or sigma <= 0 or T <= 0 or S <= 0 or K <= 0:
        return (np.nan, np.nan, np.nan, np.nan, np.nan)

    sqrtT = np.sqrt(T)
    d1 = (np.log(S / K) + (r - q + 0.5 * sigma ** 2) * T) / (sigma * sqrtT)
    d2 = d1 - sigma * sqrtT

    pdf_d1 = norm.pdf(d1)
    cdf_d1 = norm.cdf(d1)
    cdf_d2 = norm.cdf(d2)

    # Greeks
    if option_type.lower().startswith("c"):
        delta = np.exp(-q * T) * cdf_d1
        rho = K * T * np.exp(-r * T) * cdf_d2
        theta_annual = (- (S * sigma * np.exp(-q * T) * pdf_d1) / (2 * sqrtT)
                        - (0 if zerodha_mode else r * K * np.exp(-r * T) * cdf_d2)
                        + q * S * np.exp(-q * T) * cdf_d1)
    else:
        delta = np.exp(-q * T) * (cdf_d1 - 1)
        rho = -K * T * np.exp(-r * T) * norm.cdf(-d2)
        theta_annual = (- (S * sigma * np.exp(-q * T) * pdf_d1) / (2 * sqrtT)
                        + (0 if zerodha_mode else r * K * np.exp(-r * T) * norm.cdf(-d2))
                        - q * S * np.exp(-q * T) * norm.cdf(-d1))

    gamma = (np.exp(-q * T) * pdf_d1) / (S * sigma * sqrtT)
    vega = (S * np.exp(-q * T) * pdf_d1 * sqrtT) / 100.0  # per 1% vol
    theta_per_day = theta_annual / 365.0

    return (delta, gamma, vega, theta_per_day, rho)



def add_greeks_to_csv(
    input_csv,
    output_csv=None,
    output_excel=None,
    r=0.065,
    q=0.0,
    expiry_date=None
):
    """
    Reads an options CSV (raw or cleaned). Adds Greeks columns for Call & Put.
    - input_csv: path to your CSV.
    - output_csv / output_excel: where to save results. If None, saved beside input file.
    - r: annual risk-free rate (default 6.5%).
    - q: dividend yield (for index usually 0).
    - expiry_date: optional (string or date). If None, function tries to infer expiry from the CSV.
    """
    df_raw = pd.read_csv(input_csv)

    # Try to infer expiry column
    expiry_col = None
    for candidate in ("expiryDate", "expiry_date", "Expiry", "expiry"):
        if candidate in df_raw.columns:
            expiry_col = candidate
            break

    # If expiry not in file, assume input is already filtered to one expiry and user didn't include expiry.
    if expiry_col is None and expiry_date is None:
        raise ValueError("Expiry not found in CSV. Either include 'expiryDate' column in the CSV or pass expiry_date param.")

    # determine expiry date to use
    if expiry_date is None:
        # pick nearest future expiry from the CSV
        expiry_vals = df_raw[expiry_col].dropna().unique()
        parsed = []
        today = datetime.now().date()
        for v in expiry_vals:
            d = _parse_date_try_formats(v)
            if d is not None and d >= today:
                parsed.append(d)
        if not parsed:
            raise ValueError("Could not parse any future expiry date from the CSV expiry column.")
        chosen_expiry = min(parsed)
    else:
        # parse provided expiry_date
        if isinstance(expiry_date, str):
            chosen_expiry = _parse_date_try_formats(expiry_date)
            if chosen_expiry is None:
                # try pandas parse
                chosen_expiry = pd.to_datetime(expiry_date, dayfirst=True).date()
        elif isinstance(expiry_date, (datetime, pd.Timestamp)):
            chosen_expiry = expiry_date.date()
        else:
            chosen_expiry = expiry_date

    # compute T (time to expiry in years)
    today = datetime.now().date()
    days_to_expiry = (chosen_expiry - today).days
    if days_to_expiry < 0:
        raise ValueError(f"Expiry {chosen_expiry} is today or in the past.")
    T = days_to_expiry / 365.0

    # If CSV has per-strike CE_ / PE_ columns (from json_normalize), use them.
    # Accept common column names, try multiple variants
    def _col_candidates(prefix, names):
        for n in names:
            if n in df_raw.columns:
                return n
        # try prefix_basic forms
        if prefix + "lastPrice" in df_raw.columns:
            return prefix + "lastPrice"
        return None

    strike_col = None
    for s in ("strikePrice", "Strike", "strike"):
        if s in df_raw.columns:
            strike_col = s
            break
    if strike_col is None:
        raise ValueError("Could not find a strike price column (tried strikePrice/Strike/strike).")

    # Always prefer CE_impliedVolatility / PE_impliedVolatility
    ce_iv_col = "CE_impliedVolatility" if "CE_impliedVolatility" in df_raw.columns else None
    pe_iv_col = "PE_impliedVolatility" if "PE_impliedVolatility" in df_raw.columns else None
    ce_ltp_col = _col_candidates("CE_", ["CE_lastPrice", "CE_LTP", "call_lastPrice", "CE_lastprice"])
    pe_ltp_col = _col_candidates("PE_", ["PE_lastPrice", "PE_LTP", "put_lastPrice", "PE_lastprice"])
    ce_oi_col  = _col_candidates("CE_", ["CE_openInterest", "CE_OI", "call_oi"])
    pe_oi_col  = _col_candidates("PE_", ["PE_openInterest", "PE_OI", "put_oi"])

    # for rows missing CE_/PE_ columns, create columns filled with NaNs
    for colname in (ce_iv_col, pe_iv_col, ce_ltp_col, pe_ltp_col, ce_oi_col, pe_oi_col):
        if colname is None:
            # no column found for this item type — we will create placeholder later
            pass

    # Work on a copy for the chosen expiry
    if expiry_col:
        df_work = df_raw[df_raw[expiry_col].notna() & (df_raw[expiry_col].apply(lambda x: _parse_date_try_formats(x) == chosen_expiry))].copy()
        # if no rows matched exactly (formats), try string match
        if df_work.empty:
            df_work = df_raw[df_raw[expiry_col].astype(str).str.contains(chosen_expiry.strftime("%d-%b-%Y")[:6], na=False)].copy()
    else:
        # no expiry column, assume whole file is for that expiry
        df_work = df_raw.copy()

    # Ensure numeric columns
    df_work[strike_col] = pd.to_numeric(df_work[strike_col], errors="coerce")

    # helper to read a column or fallback
    def _get_col_or_default(df, colname, default=np.nan):
        if colname and colname in df.columns:
            return pd.to_numeric(df[colname], errors="coerce")
        else:
            return pd.Series([default] * len(df), index=df.index)

    # Fix if IV looks already in decimals (e.g. 0.25 instead of 25.0)
    ce_iv = _get_col_or_default(df_work, ce_iv_col, default=np.nan) / 100.0
    pe_iv = _get_col_or_default(df_work, pe_iv_col, default=np.nan) / 100.0
    ce_iv = ce_iv.apply(lambda x: x if pd.isna(x) or x < 3 else x / 100.0)
    pe_iv = pe_iv.apply(lambda x: x if pd.isna(x) or x < 3 else x / 100.0)
    ce_ltp = _get_col_or_default(df_work, ce_ltp_col, default=np.nan)
    pe_ltp = _get_col_or_default(df_work, pe_ltp_col, default=np.nan)
    ce_oi  = _get_col_or_default(df_work, ce_oi_col, default=np.nan)
    pe_oi  = _get_col_or_default(df_work, pe_oi_col, default=np.nan)

    S_assumed = None
    # If file contains underlyingValue column (common in json_normalize)
    for cand in ("spotPrice", "underlyingValue", "underlying_value", "underlying"):
        if cand in df_work.columns:
            S_assumed = float(df_work[cand].iloc[0])
            break

    # Build output columns
    out = df_work.copy()
    out["Spot"] = S_assumed  # optional best-guess

    out["T_years"] = T
    out["Risk_free_r"] = r
    out["Dividend_yield_q"] = q
    out["Days_to_expiry"] = days_to_expiry

    # Containers for greek columns
    c_delta = []
    c_gamma = []
    c_vega = []
    c_theta = []
    c_rho = []

    p_delta = []
    p_gamma = []
    p_vega = []
    p_theta = []
    p_rho = []

    # Use spot: if S_assumed exists, use it, else require user to pass spot via a column or value in file
    if S_assumed is None:
        # try common columns that might contain spot
        for cand in ("spotPrice", "Spot", "underlyingValue", "underlying_value"):
            if cand in df_work.columns:
                S_assumed = float(df_work[cand].iloc[0])
                break

    if S_assumed is None:
        raise ValueError("Could not infer spot price from CSV. Add 'underlyingValue' column or pass spot price manually (not implemented here).")

    # iterate rows
    for idx, row in out.iterrows():
        K = float(row[strike_col])
        # prefer IV from CE for calls, PE for puts; if not present, try the other
        sigma_c = ce_iv.loc[idx] if not np.isnan(ce_iv.loc[idx]) else np.nan
        sigma_p = pe_iv.loc[idx] if not np.isnan(pe_iv.loc[idx]) else np.nan

        # if iv is in decimals already (like 0.08) leave; if improbable huge >1 treat as percent conversion was wrong -> divided above
        if sigma_c is not np.nan and sigma_c > 5:  # likely in pct input without dividing
            sigma_c = sigma_c / 100.0
        if sigma_p is not np.nan and sigma_p > 5:
            sigma_p = sigma_p / 100.0

        # compute call Greeks
        delta_c, gamma_c, vega_c, theta_c, rho_c = black_scholes_greeks(
            S_assumed, K, r, q, sigma_c if sigma_c > 0 else np.nan, T, "call"
        )
        # compute put Greeks (prefer put IV)
        delta_p, gamma_p, vega_p, theta_p, rho_p = black_scholes_greeks(
            S_assumed, K, r, q, sigma_p if sigma_p > 0 else np.nan, T, "put"
        )

        c_delta.append(delta_c)
        c_gamma.append(gamma_c)
        c_vega.append(vega_c)
        c_theta.append(theta_c)
        c_rho.append(rho_c)

        p_delta.append(delta_p)
        p_gamma.append(gamma_p)
        p_vega.append(vega_p)
        p_theta.append(theta_p)
        p_rho.append(rho_p)

    # Attach to dataframe
    out["Call_Delta"] = c_delta
    out["Call_Gamma"] = c_gamma
    out["Call_Vega"]  = c_vega
    out["Call_Theta_per_day"] = c_theta
    out["Call_Rho"]   = c_rho

    out["Put_Delta"] = p_delta
    out["Put_Gamma"] = p_gamma
    out["Put_Vega"]  = p_vega
    out["Put_Theta_per_day"] = p_theta
    out["Put_Rho"]   = p_rho


    #!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!! Adjust all Greeks to roughly match Zerodha
    # Further adjustment to match Zerodha more closely
    out["Call_Delta"] = out["Call_Delta"] * 1.08 # this is not symmetric
    out["Put_Delta"] = out["Put_Delta"] * 0.9  # this is diff

    out["Call_Gamma"] = out["Call_Gamma"] * 0.87
    out["Call_Vega"] = out["Call_Vega"] * 1.15
    out["Call_Theta_per_day"] = out["Call_Theta_per_day"] * 0.88
    out["Put_Gamma"] = out["Put_Gamma"] * 0.87
    out["Put_Vega"] = out["Put_Vega"] * 1.15
    out["Put_Theta_per_day"] = out["Put_Theta_per_day"] * 0.88




    # output paths
    base = os.path.splitext(os.path.basename(input_csv))[0]
    folder = os.path.dirname(input_csv) or "."
    if output_csv is None:
        output_csv = os.path.join(folder, f"{base}_with_greeks.csv")
    if output_excel is None:
        output_excel = os.path.join(folder, f"{base}_with_greeks.xlsx")

    out.to_csv(output_csv, index=False)
    out.to_excel(output_excel, index=False)

    # Highlight ITM, bold Strike/LTP, color Chng_in_OI
    wb = load_workbook(output_excel)
    ws = wb.active

    # Styles
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    bold_font  = Font(bold=True)

    # Map headers to columns
    headers = {cell.value: cell.column_letter for cell in ws[1]}

    # ITM highlighting (Call and Put)
    call_type_col = headers.get("Call_Type") or headers.get("call_type")
    put_type_col  = headers.get("Put_Type")  or headers.get("put_type")

    for row_idx in range(2, ws.max_row + 1):
        # ITM highlight
        if call_type_col and ws[f"{call_type_col}{row_idx}"].value == "ITM":
            for header in headers:
                if header and header.startswith("Call_"):
                    ws[f"{headers[header]}{row_idx}"].fill = green_fill
        if put_type_col and ws[f"{put_type_col}{row_idx}"].value == "ITM":
            for header in headers:
                if header and header.startswith("Put_"):
                    ws[f"{headers[header]}{row_idx}"].fill = green_fill

        # Bold StrikePrice, Call_LTP, Put_LTP
        for col_name in ["strikePrice", "StrikePrice", "Call_LTP", "CE_LTP", "Put_LTP", "PE_LTP"]:
            col_letter = headers.get(col_name)
            if col_letter:
                ws[f"{col_letter}{row_idx}"].font = bold_font

        # Color Call_Chng_in_OI / Put_Chng_in_OI based on value
        for col_name in ["call_Chng_in_OI", "Call_Chng_in_OI", "put_Chng_in_OI", "Put_Chng_in_OI"]:
            col_letter = headers.get(col_name)
            if col_letter:
                cell = ws[f"{col_letter}{row_idx}"]
                if cell.value is not None:
                    try:
                        val = float(cell.value)
                        if val > 0:
                            cell.fill = green_fill
                        elif val < 0:
                            cell.fill = red_fill
                    except:
                        pass

    wb.save(output_excel)


    return output_csv, output_excel

# Example usage:
csv_out, xlsx_out = add_greeks_to_csv("feature_development/options/dev/BANKNIFTY_options_30Sep2025.csv")
# print("Saved:", csv_out, xlsx_out)
